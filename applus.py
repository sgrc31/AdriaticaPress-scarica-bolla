#!/usr/bin/env python3

import sys
import time
import csv
import os
from PyQt5.QtWidgets import QWidget, QLabel, QApplication, QPushButton, QCalendarWidget, QHBoxLayout, QVBoxLayout, QDialog, QTextEdit, QFileDialog, QMessageBox
#from PyQt5.QtCore import pyqtSlot
from PyQt5.QtCore import QDate
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException, NoSuchElementException
from openpyxl import load_workbook

ip_ap = 'http://46.37.234.75/'
if os.name == 'nt':
    PATH_BOLLA_SCARICATA = 'C:\\EasyRetail\\Carico\\Giornali\\bolla.csv'
    PATH_FILE_VENDITE = os.path.expanduser('~/Desktop/vendite_giornali/Foglio1.xlsx')
else:
    PATH_BOLLA_SCARICATA = 'bolla.csv'
    PATH_FILE_VENDITE = 'Foglio1.xlsx'

class MyWin(QDialog):
    
    def __init__(self, parent=None):
        super(MyWin, self).__init__(parent)
        self.initUI()
        
    def initUI(self):
        self.calendar_label = QLabel('Seleziona la data', self)
        self.calendar = QCalendarWidget(self)
        self.stringa_data = time.strftime('%d/%m/%Y') # per far si che ci sia un valore all'avvio, da inserire con strftime
        self.stringa_data_per_nome_file = time.strftime('%d%m%y')
        #Creo i bottoni
        self.btn_download = QPushButton('Scarica Bolle', self)
        self.btn_upload = QPushButton('Carica Vendite', self)
        #Creo widget testo per output
        self.mio_testo = QTextEdit(self)
        self.mio_testo.setReadOnly(True)
        self.mio_testo.setMaximumHeight(50)
        #Creo layout
        self.vbox = QVBoxLayout()
        self.hbox = QHBoxLayout()
        self.hbox.addWidget(self.btn_download)
        self.hbox.addWidget(self.btn_upload)
        self.vbox.addWidget(self.calendar_label)
        self.vbox.addWidget(self.calendar)
        self.vbox.addLayout(self.hbox)
        self.vbox.addWidget(self.mio_testo)
        self.setLayout(self.vbox)
        self.calendar.clicked[QDate].connect(self.set_date)
        #self.btn_download.clicked.connect(lambda: self.test_funct('ciccio')) #uso lambda per passare ulteriori argomenti al segnale
        self.btn_download.clicked.connect(self.download_bolle)
        self.btn_upload.clicked.connect(self.upload_vendite)
        self.setWindowTitle('Occupy Artoni')
        self.show()

    def set_date(self, date):
        '''Imposta i valori iniziali per la data, tanto la stringa da passare come url
        che la stringa derivata da utilizzare come parte del nome output
        '''
        self.stringa_data = date.toString('dd/MM/yyyy')
        self.stringa_data_per_nome_file = date.toString('ddMMyy')

    def download_bolle(self):
        self.login_adriaticapress()
        self.scarica_bolle()
        self.destroy_webdriwer()

    def upload_vendite(self):
        if os.path.isfile(PATH_FILE_VENDITE):
            self.login_adriaticapress()
            self.inserimento_vendite()
            self.destroy_webdriwer()
        else:
            return self.mio_testo.setPlainText('Nessun file vendite trovato\nAssicurarsi di aver fatto export delle vendite da EasyRetail')

    def login_adriaticapress(self):
        '''Creazione webdriver e procedura di login in adriatipress
        '''
        try:
            with open('secrets.txt') as credentials_file:
                credentials_file_list = list(credentials_file.readlines())
                username = credentials_file_list[0].strip()[9:]
                password = credentials_file_list[1].strip()[9:]
        except FileNotFoundError:
            self.msg = QMessageBox()
            self.msg.setText('File credenziali non presente')
            self.msg.setWindowTitle('testtitolo')
            self.msg.setStandardButtons(QMessageBox.Ok)
            return self.msg.exec_()
        if os.name == 'nt':
            self.dr = webdriver.Chrome('chromedriver.exe')
        else:
            self.dr = webdriver.Chrome()
        self.dr.maximize_window()
        self.dr.get('{}Login.htm'.format(ip_ap))
        time.sleep(5)
        username_field = WebDriverWait(self.dr, 15).until(EC.presence_of_element_located((By.ID, 'txtUsername')))
        password_field = WebDriverWait(self.dr, 15).until(EC.presence_of_element_located((By.ID, 'txtPassword')))
        username_field.clear()
        username_field.send_keys(username)
        password_field.clear()
        password_field.send_keys(password)
        password_field.send_keys(Keys.RETURN)
        time.sleep(5)


    def scarica_bolle(self):
        '''Gestisce l'effettivo dowload di entrambi i tipi di bolla,
        con handling di eventuali errori dati da mancaza di bolla o
        timeout della connessione
        '''
        #Scarico prima bolla
        self.dr.get('{}Bolla.htm-{}-B'.format(ip_ap, self.stringa_data))
        try:
            if self.dr.title == 'Errore di runtime':
                raise ValueError('testo raise')
            self.scarica_bolla('B')
        except WebDriverException:
            self.mio_testo.insertPlainText('Errore imprevisto\n')
        except ValueError:
            self.mio_testo.insertPlainText('{} Bolla tipo B non presente\n'.format(time.strftime('%H:%M')))
        #Scarico seconda bolla
        #self.dr.get('{}Bolla.htm-{}-C'.format(ip_ap, self.stringa_data))
        #try:
        #    if self.dr.title == 'Errore di runtime':
        #        raise ValueError('testo raise')
        #    self.scarica_bolla('C')
        #except WebDriverException:
        #    self.mio_testo.insertPlainText('errore imprevisto\n')
        #except ValueError:
        #    self.mio_testo.insertPlainText('{} Bolla tipo C non presente\n'.format(time.strftime('%H:%M')))

    def scarica_bolla(self, tipo_bolla):
        '''Procedura di raccolta dati bolla dal portale adriaticapress e scrittura
        su file csv. tipo_bolla può essere uno dei 2 valori ['B', 'C']
        '''
        #Inserisco elemento di controllo sul caricamento della bolla vera e propria
        elemento_controllo = WebDriverWait(self.dr, 300).until(EC.invisibility_of_element_located((By.ID, 'loaderBodyPage')))
        time.sleep(1)
        #Rendo header statico, così da non interferier nello scroll
        self.dr.execute_script('$(".superHeader").css({position: "static"});')
        lista_bolla = []
        elenco_link_testate = self.dr.find_elements_by_css_selector('a[id^=\"ctl00_MainContent_dlBolla_ctl\"]')
        for link in elenco_link_testate:
            dati_testata = []
            #Assicurarsi che il link sia visibile prima di clickarlo
            self.dr.execute_script('arguments[0].scrollIntoView();', link)
            link.click()
            nome_testata = WebDriverWait(self.dr, 300).until(EC.presence_of_element_located((By.ID, 'lblTitoloDettaglio'))).text
            identificativo_testata = self.dr.find_element_by_id('lblCodice').text
            numero_testata = self.dr.find_element_by_id('lblNumeroDettaglio').text
            prezzo_testata = self.dr.find_element_by_id('lblPrezzoDettaglio').text[2:].replace(',', '.')
            try:
                barcode_testata = self.dr.find_element_by_id('lblBarcodeDettaglio').text
            except NoSuchElementException:
                barcode_testata = 'non presente'
            dati_testata.extend([nome_testata, identificativo_testata, numero_testata, barcode_testata, prezzo_testata])
            lista_bolla.append(dati_testata)
        with open(PATH_BOLLA_SCARICATA, 'w') as output_file:
            output_writer = csv.writer(output_file)
            for row in lista_bolla:
                output_writer.writerow(row)
        return self.mio_testo.insertPlainText('{} File {} creato\n'.format(time.strftime('%H:%M'), output_file.name))

    def inserimento_vendite(self):
        '''Procedura di immissione delle vendite su adriaticapress
        '''
        self.dr.get('{}Vendita.htm'.format(ip_ap))
        time.sleep(5)
        elemento_controllo = WebDriverWait(self.dr, 300).until(EC.invisibility_of_element_located((By.ID, 'loaderBodyPage')))
        time.sleep(5)
        self.dr.find_element_by_id('imgMostraRicerca').click()
        input_ean = self.dr.find_element_by_name('txtTitoloRI')
        input_qtt = self.dr.find_element_by_name('txtQ')
        wb = load_workbook(PATH_FILE_VENDITE)
        ws = wb.get_sheet_by_name('ExportExcel')
        righe_con_dati = ws.max_row - 1
        for riga in ws.iter_rows(min_row=3, max_row=righe_con_dati):
            descrizione, ean, copie = riga
            time.sleep(1)
            input_ean.clear()
            input_ean.send_keys(ean.value)
            input_qtt.clear()
            input_qtt.send_keys(copie.value)
            input_ean.send_keys(Keys.RETURN)
        os.remove(PATH_FILE_VENDITE)
        return self.mio_testo.insertPlainText('{} Vendite caricate\n'.format(time.strftime('%H:%M')))

    def destroy_webdriwer(self):
        '''Distrugge il webdriver al termine delle operazioni e
        ne dà comunicazione all'utente
        '''
        self.dr.quit()
        return self.mio_testo.insertPlainText('{} Procedura terminata\n'.format(time.strftime('%H:%M')))



    #def carica_vendite(self):
    #    self.file_vendite = QFileDialog.getOpenFileName(self, 'Seleziona file vendite', os.path.abspath(os.path.dirname('.')))[0]
    #    #Per selezionare una cartella invece che un file
    #    #dirname = QFileDialog.getExistingDirectory(None, 'Select a folder:', '/', QFileDialog.ShowDirsOnly)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyWin()
    sys.exit(app.exec_())
